import os
import json
import csv
import io
import re
import uuid
import datetime

import memory as mem

# app_settings 테이블(Turso/SQLite, memory.save_setting·get_setting 경유)에 저장.
# 과거에는 로컬 JSON 파일(budget-data.json)에 저장했으나, 컨테이너 재배포 시
# 파일시스템이 초기화되는 배포 환경(Render 등)에서 데이터가 유실되는 문제가 있어
# 대화·학습 데이터와 동일하게 DB에 영속 저장하도록 변경.
_BUDGET_KEY = "budget_data"
MONTHS = list(range(1, 13))
CATEGORIES = ["판관", "용역", "경상"]


def _empty():
    return {"headcount": [], "items": [], "uploads": [], "grid": [], "snapshots": []}


def _now_iso():
    return datetime.datetime.now().isoformat(timespec="seconds")


def read_budget():
    data = mem.get_setting(_BUDGET_KEY, None)
    if data is None:
        return _empty()
    for key, default in _empty().items():
        if key not in data:
            data[key] = default
    return data


def write_budget(data):
    mem.save_setting(_BUDGET_KEY, data)


def to_number(v):
    if v is None:
        return None
    s = str(v).strip().replace(",", "")
    if s == "":
        return None
    try:
        return float(s) if "." in s else int(s)
    except ValueError:
        return None


# 파싱 행 수 상한. 엑셀은 시트 하나에 100만 행을 담을 수 있어, 크기 제한(MB)만으로는
# 압축률이 높은 파일이 메모리에서 수십 배로 부풀 수 있다(zip bomb 유사). 실사용
# 예산 시트 규모를 크게 웃도는 값으로 잡되 무제한은 막는다.
MAX_PARSE_ROWS = int(os.getenv("MAX_PARSE_ROWS", "50000"))


class TooManyRowsError(Exception):
    """업로드 시트의 행 수가 상한을 넘었을 때"""


def _guard_rows(n):
    if n > MAX_PARSE_ROWS:
        raise TooManyRowsError(
            f"업로드 시트의 행 수가 너무 많습니다(최대 {MAX_PARSE_ROWS:,}행). "
            f"파일을 나누거나 MAX_PARSE_ROWS 환경변수를 조정하세요."
        )


def parse_rows(content: bytes, filename: str):
    """업로드 파일(xlsx/csv)을 헤더가 있는 dict 행 리스트로 변환"""
    if filename.lower().endswith((".xlsx", ".xls")):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        sheet = wb.worksheets[0]
        rows_iter = sheet.iter_rows(values_only=True)
        header = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
        rows = []
        for r in rows_iter:
            rows.append({header[i]: r[i] for i in range(len(header)) if i < len(r)})
            _guard_rows(len(rows))
        return rows
    else:
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        rows = []
        for row in reader:
            rows.append(dict(row))
            _guard_rows(len(rows))
        return rows


def upsert_headcount(rows):
    data = read_budget()
    upserted = 0
    depts = set()

    for row in rows:
        dept = row.get("구분")
        if not dept or dept == "계":
            continue
        depts.add(dept)

        for m in MONTHS:
            value = to_number(row.get(f"{m}월"))
            if value is None:
                continue
            existing = next((h for h in data["headcount"] if h["dept"] == dept and h["month"] == m), None)
            if existing:
                existing["count"] = value
            else:
                data["headcount"].append({"dept": dept, "month": m, "count": value})
            upserted += 1

    data["uploads"].append({"type": "headcount", "rows": len(rows)})
    write_budget(data)
    return upserted, sorted(depts)


def upsert_detail(rows):
    data = read_budget()
    upserted = 0
    depts = set()

    for row in rows:
        dept = row.get("부문")
        category = row.get("구분")
        if not dept or dept == "계" or category not in CATEGORIES:
            continue
        depts.add(dept)

        team = row.get("팀") or ""
        revenue_type = row.get("매출구분") or ""
        account = row.get("항목") or ""
        detail = row.get("세부내역(산정근거)") or row.get("세부내역") or ""

        for m in MONTHS:
            amount = to_number(row.get(f"{m}월"))
            if amount is None:
                continue
            existing = next((
                i for i in data["items"]
                if i["dept"] == dept and i["team"] == team and i["account"] == account
                and i["category"] == category and i["month"] == m
            ), None)
            if existing:
                existing["amount"] = amount
                existing["revenueType"] = revenue_type
                existing["detail"] = detail
            else:
                data["items"].append({
                    "dept": dept, "team": team, "revenueType": revenue_type,
                    "account": account, "detail": detail, "category": category,
                    "month": m, "amount": amount
                })
            upserted += 1

    data["uploads"].append({"type": "detail", "rows": len(rows)})
    write_budget(data)
    return upserted, sorted(depts)


def build_summary():
    data = read_budget()
    depts = sorted(set([h["dept"] for h in data["headcount"]] + [i["dept"] for i in data["items"]]))

    summary = []
    for dept in depts:
        months = []
        for m in MONTHS:
            headcount_entry = next((h for h in data["headcount"] if h["dept"] == dept and h["month"] == m), None)
            dept_items = [i for i in data["items"] if i["dept"] == dept and i["month"] == m]

            by_category = {c: 0 for c in CATEGORIES}
            for i in dept_items:
                by_category[i["category"]] += i["amount"]

            total_amount = sum(i["amount"] for i in dept_items)

            months.append({
                "month": m,
                "headcount": headcount_entry["count"] if headcount_entry else None,
                **by_category,
                "totalAmount": total_amount,
                "hasHeadcountData": headcount_entry is not None,
                "hasDetailData": len(dept_items) > 0
            })
        summary.append({"dept": dept, "months": months})

    return summary


# ------------------- 엑셀형 그리드 (예산 시트) -------------------

def get_grid():
    data = read_budget()
    return data.get("grid") or []


def save_grid(rows):
    data = read_budget()
    data["grid"] = rows
    data["grid_updated_at"] = _now_iso()
    write_budget(data)
    return True


def get_sheets():
    """멀티시트 데이터 반환 — {sheets, active, styles}"""
    data = read_budget()
    if "sheets" in data:
        return {
            "sheets": data["sheets"],
            "active": data.get("active", list(data["sheets"].keys())[0]),
            "styles": data.get("styles", {}),
            "colWidths": data.get("colWidths", {}),
            "rowHeights": data.get("rowHeights", {}),
            "mergedCells": data.get("mergedCells", {}),
            "conditionalFormats": data.get("conditionalFormats", {}),
            "dataValidations": data.get("dataValidations", {}),
            "cellNotes": data.get("cellNotes", {}),
        }
    # 기존 단일 grid → Sheet1으로 마이그레이션
    grid = data.get("grid") or []
    return {"sheets": {"Sheet1": grid}, "active": "Sheet1", "styles": {}, "colWidths": {}, "rowHeights": {}}


def save_sheets(payload: dict):
    """멀티시트 데이터 저장 — payload: {sheets, active, styles, colWidths, rowHeights}"""
    data = read_budget()
    data["sheets"] = payload.get("sheets", {})
    data["active"] = payload.get("active", "Sheet1")
    data["styles"] = payload.get("styles", {})
    data["colWidths"] = payload.get("colWidths", {})
    data["rowHeights"] = payload.get("rowHeights", {})
    data["mergedCells"] = payload.get("mergedCells", {})
    data["conditionalFormats"] = payload.get("conditionalFormats", {})
    data["dataValidations"] = payload.get("dataValidations", {})
    data["cellNotes"] = payload.get("cellNotes", {})
    data["sheets_updated_at"] = _now_iso()
    # 하위 호환: 활성 시트를 grid에도 동기
    active = data["active"]
    if active in data["sheets"]:
        data["grid"] = data["sheets"][active]
    write_budget(data)
    return True


def parse_grid(content: bytes, filename: str):
    """업로드 파일(xlsx/csv)을 그대로 2차원 문자열 배열(grid)로 변환"""
    if filename.lower().endswith((".xlsx", ".xls")):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        sheet = wb.worksheets[0]
        rows = []
        for r in sheet.iter_rows(values_only=True):
            rows.append(["" if v is None else str(v) for v in r])
            _guard_rows(len(rows))
        return rows
    else:
        text = content.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text))
        rows = []
        for row in reader:
            rows.append(list(row))
            _guard_rows(len(rows))
        return rows


# static/budget.html의 자체 수식 엔진(FUNCS 객체)이 실제로 구현·지원하는 함수
# 이름 전체. 이 목록은 화이트리스트로만 쓰인다 — 여기 없는 함수 호출은
# (문법이 그럴듯해 보여도) 안전하지 않은 것으로 간주해 텍스트로 다운그레이드한다.
# static/budget.html 3478~3679번째 줄 부근 `const FUNCS = {...}` 객체의 키와
# 반드시 동기화되어야 함 — 그쪽에 함수를 추가하면 여기도 함께 갱신할 것.
_SAFE_FORMULA_FUNCS = {
    "ABS", "AND", "AVERAGE", "AVERAGEIF", "AVERAGEIFS", "CEILING", "CONCAT",
    "CONCATENATE", "COUNT", "COUNTA", "COUNTIF", "COUNTIFS", "DAY", "FLOOR",
    "HLOOKUP", "IF", "IFERROR", "IFNA", "IFS", "INDEX", "INT", "ISBLANK",
    "ISERROR", "ISNUMBER", "LARGE", "LEFT", "LEN", "LOWER", "MATCH", "MAX",
    "MEDIAN", "MID", "MIN", "MOD", "MONTH", "NOT", "NOW", "OR", "POWER",
    "RANK", "RIGHT", "ROUND", "ROUNDDOWN", "ROUNDUP", "SMALL", "SQRT",
    "STDEV", "SUM", "SUMIF", "SUMIFS", "SUMPRODUCT", "TEXT", "TODAY",
    "TRIM", "TRUNC", "UPPER", "VALUE", "VLOOKUP", "YEAR",
}
# 함수 호출처럼 보이는 토큰(식별자 바로 뒤에 여는 괄호)을 찾는 패턴.
_FORMULA_FUNC_CALL_RE = re.compile(r"[A-Za-z_][A-Za-z0-9_]*(?=\s*\()")
# 파이프(|)는 이 앱의 수식 문법에 전혀 쓰이지 않는 문자이면서, 동시에
# `=cmd|'/c calc'!A1` 류의 고전적인 Excel DDE/명령실행 인젝션 페이로드의
# 핵심 구성요소다 — 화이트리스트를 통과하는 함수명이 하나도 없는 수식이라도
# 이 문자가 섞여 있으면 무조건 안전하지 않은 것으로 취급한다(심층 방어).
_FORMULA_DANGEROUS_CHAR_RE = re.compile(r"[|]")


def _is_safe_export_formula(val: str) -> bool:
    """이 그리드의 자체 수식 엔진이 실제로 파싱·계산하는 함수만 호출하는
    수식인지 화이트리스트 기준으로 판정한다(블랙리스트가 아니라 화이트리스트 —
    미지의/신규 위험 함수를 놓칠 걱정이 없다). HYPERLINK·WEBSERVICE처럼 외부
    실행·네트워크를 유발하는 실제 Excel 함수는 이 화이트리스트에 없으므로
    항상 안전하지 않음으로 판정된다."""
    if _FORMULA_DANGEROUS_CHAR_RE.search(val):
        return False
    for m in _FORMULA_FUNC_CALL_RE.finditer(val):
        if m.group(0).upper() not in _SAFE_FORMULA_FUNCS:
            return False
    return True


def export_grid_xlsx(rows):
    import openpyxl
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "예산"
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, val in enumerate(row, start=1):
            cell = sheet.cell(row=r_idx, column=c_idx)
            if isinstance(val, str) and val.startswith("=") and _is_safe_export_formula(val):
                cell.value = val
            elif isinstance(val, str) and val.startswith("="):
                # 이 그리드가 지원하지 않는 함수·위험 문자(| 등)가 섞인 값은
                # 실제 Excel 수식으로 절대 내보내지 않는다(Formula/CSV Injection
                # 방지 — 다른 사람이 실제 Microsoft Excel에서 이 파일을 열었을 때
                # 임의 명령 실행·외부 URL 호출·클릭유도 링크가 생기는 것을 막기
                # 위함). openpyxl은 "=" 로 시작하는 문자열을 값으로 대입하면
                # 자동으로 수식 셀(data_type='f')로 인식하므로, 대입 후 명시적으로
                # data_type을 문자열('s')로 되돌려 원문 그대로를 "글자"로만 저장한다.
                cell.value = val
                cell.data_type = "s"
            else:
                num = to_number(val)
                cell.value = num if num is not None else val
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ------------------- 스냅샷(월별 최종 자료) -------------------

def list_snapshots():
    data = read_budget()
    return [
        {"id": s["id"], "label": s["label"], "createdAt": s["createdAt"], "rows": len(s["grid"])}
        for s in data.get("snapshots", [])
    ]


def save_snapshot(label, rows):
    data = read_budget()
    snapshots = data.setdefault("snapshots", [])
    existing = next((s for s in snapshots if s["label"] == label), None)
    if existing:
        existing["grid"] = rows
        existing["createdAt"] = _now_iso()
        snap_id = existing["id"]
    else:
        snap_id = str(uuid.uuid4())
        snapshots.append({
            "id": snap_id, "label": label, "grid": rows, "createdAt": _now_iso()
        })
    write_budget(data)
    return snap_id


def get_snapshot(snap_id):
    data = read_budget()
    return next((s for s in data.get("snapshots", []) if s["id"] == snap_id), None)


def delete_snapshot(snap_id):
    data = read_budget()
    before = len(data.get("snapshots", []))
    data["snapshots"] = [s for s in data.get("snapshots", []) if s["id"] != snap_id]
    write_budget(data)
    return before != len(data["snapshots"])


_CELL_RE = re.compile(r"^[A-Za-z]+\d+$")


def _col_letters(idx):
    """0-based column index -> Excel-style letters"""
    letters = ""
    idx += 1
    while idx > 0:
        idx, rem = divmod(idx - 1, 26)
        letters = chr(65 + rem) + letters
    return letters


def compare_rows(rows_a, rows_b):
    """두 그리드(rows_a=기준, rows_b=비교대상)를 셀 단위로 비교, 숫자면 차이 계산"""
    n_rows = max(len(rows_a), len(rows_b))
    diffs = []
    for r in range(n_rows):
        row_a = rows_a[r] if r < len(rows_a) else []
        row_b = rows_b[r] if r < len(rows_b) else []
        n_cols = max(len(row_a), len(row_b))
        for c in range(n_cols):
            val_a = row_a[c] if c < len(row_a) else ""
            val_b = row_b[c] if c < len(row_b) else ""
            val_a = "" if val_a is None else str(val_a)
            val_b = "" if val_b is None else str(val_b)
            if val_a == val_b:
                continue
            num_a = to_number(val_a)
            num_b = to_number(val_b)
            delta = None
            if num_a is not None and num_b is not None:
                delta = num_b - num_a
            diffs.append({
                "cell": f"{_col_letters(c)}{r + 1}",
                "row": r, "col": c,
                "before": val_a, "after": val_b, "delta": delta
            })
    return diffs
